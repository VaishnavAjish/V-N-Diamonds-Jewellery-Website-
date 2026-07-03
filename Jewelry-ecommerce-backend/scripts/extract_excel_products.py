"""
Reads a VN Diamonds-style vendor Excel sheet, maps its columns to our product
fields, and extracts each row's "Photo" image to a file on disk, keyed by the
"Lot Number" (SKU).

Expected header row (in this order, "Metal" appears twice):
    Jewellery | Product | Poetic Name | Lot Number | Description | Metal |
    Metal gms | Diamond Pcs | Shape | Certificate | Cert.No | Metal |
    Gross Gms | Photo | USD Price

The "Photo" column uses Excel's modern "Place in Cell" / IMAGE() rich-value
feature (not a classic floating drawing), so a plain `cell.value` read only
returns "#VALUE!". The real image bytes live in xl/media/*, linked through
xl/richData/richValueRel.xml -> xl/richData/rdrichvalue.xml -> xl/metadata.xml
-> the cell's `vm` attribute in the worksheet XML. This script resolves that
chain directly. Classic floating/anchored images (older vendor files) are
also supported as a fallback.

Usage:
    python extract_excel_products.py <xlsx_path> <image_output_dir>

Prints a single JSON array to stdout — one object per data row:
    {
        "row": <1-indexed excel row>,
        "sku": "...",            # from "Lot Number"
        "categoryName": "...",   # from "Jewellery"
        "subCategory": "...",    # from "Product"
        "settingType": "...",    # from "Poetic Name"
        "description": "...",
        "metalId": "...",        # from the first "Metal" column
        "metalGms": "...",       # from "Metal gms"
        "metalCombined": "...",  # from the second "Metal" column (e.g. "18K 5.24gms")
        "grossGms": "...",       # from "Gross Gms"
        "saleTotal": "...",      # from "USD Price"
        "diamondPcs": "...",
        "shape": "...",
        "certificate": "...",
        "certNo": "...",
        "imagePath": "<abs path to extracted image or null>"
    }
"""
import sys
import os
import json
import zipfile
import posixpath
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_XLRD = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"
NS_RVREL = "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKGREL = "http://schemas.openxmlformats.org/package/2006/relationships"

# target field -> accepted source header names (case-insensitive, trimmed)
COLUMN_ALIASES = {
    "categoryName": ["jewellery", "category"],
    "subCategory": ["product", "sub category", "subcategory"],
    "settingType": ["poetic name", "setting type"],
    "sku": ["lot number", "sku"],
    "description": ["description"],
    "metalGms": ["metal gms"],
    "diamondPcs": ["diamond pcs", "diamdon pcs"],
    "shape": ["shape"],
    "certificate": ["certificate"],
    "certNo": ["cert.no", "cert no", "certno"],
    "grossGms": ["gross gms"],
    "photo": ["photo", "image", "image file"],
    "saleTotal": ["usd price", "sale total"],
}


def normalize(header):
    return str(header).strip().lower() if header is not None else ""


def build_header_index(ws):
    """Maps field -> 0-indexed column. Handles the duplicate "Metal" header
    by assigning the first occurrence to metalId and the second to metalCombined."""
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [normalize(c.value) for c in row]
        if any(values):
            header_row = row
            break
    if header_row is None:
        raise ValueError("Could not find a header row")

    index = {}
    metal_seen = False
    for cell in header_row:
        norm = normalize(cell.value)
        col = cell.column - 1
        if norm == "metal":
            index["metalId" if not metal_seen else "metalCombined"] = col
            metal_seen = True
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if norm in aliases and field not in index:
                index[field] = col
    return index, header_row[0].row


def extract_classic_images(ws, out_dir):
    """Fallback for older files that use plain floating/anchored drawings."""
    row_to_path = {}
    for image in getattr(ws, "_images", []):
        anchor = getattr(image, "anchor", None)
        anchor_from = getattr(anchor, "_from", None) if anchor else None
        if anchor_from is None:
            continue
        excel_row = anchor_from.row + 1
        try:
            data = image._data()
        except Exception:
            continue
        ext = "." + str(getattr(image, "format", "png") or "png").lower().replace("jpeg", "jpg")
        path = os.path.join(out_dir, f"row_{excel_row}{ext}")
        with open(path, "wb") as f:
            f.write(data)
        row_to_path[excel_row] = path
    return row_to_path


def find_sheet_xml_path(zf, sheet_name):
    """Resolves the physical worksheet XML path (e.g. xl/worksheets/sheet1.xml)
    for a given sheet name, via workbook.xml + workbook.xml.rels."""
    wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {
        rel.get("Id"): rel.get("Target")
        for rel in rels_xml.findall(f"{{{NS_PKGREL}}}Relationship")
    }
    for sheet in wb_xml.findall(f"{{{NS_MAIN}}}sheets/{{{NS_MAIN}}}sheet"):
        if sheet.get("name") == sheet_name:
            rid = sheet.get(f"{{{NS_R}}}id")
            target = rid_to_target.get(rid)
            if target:
                return posixpath.normpath(posixpath.join("xl", target))
    return None


def extract_rich_value_images(zf, sheet_xml_path, photo_col_0idx, out_dir):
    """Resolves Excel's "Place in Cell" image-in-cell feature:
    cell[vm] -> metadata.xml bk[vm-1] -> rvb index -> rdrichvalue.xml rv[index]
    -> richValueRel index -> richValueRel.xml.rels -> xl/media/imageN.*"""
    required = ["xl/metadata.xml", "xl/richData/rdrichvalue.xml",
                "xl/richData/richValueRel.xml", "xl/richData/_rels/richValueRel.xml.rels"]
    if not all(name in zf.namelist() for name in required) or not sheet_xml_path:
        return {}

    # vm (1-based, in document order of <bk> under futureMetadata) -> rv index
    metadata_xml = ET.fromstring(zf.read("xl/metadata.xml"))
    bk_to_rv = []
    for bk in metadata_xml.iter(f"{{{NS_MAIN}}}bk"):
        rvb = bk.find(f".//{{{NS_XLRD}}}rvb")
        bk_to_rv.append(int(rvb.get("i")) if rvb is not None else None)

    # rv index -> richValueRel index
    rv_xml = ET.fromstring(zf.read("xl/richData/rdrichvalue.xml"))
    rv_to_relindex = []
    for rv in rv_xml.findall(f"{{{NS_XLRD}}}rv"):
        first_v = rv.find(f"{{{NS_XLRD}}}v")
        rv_to_relindex.append(int(first_v.text) if first_v is not None else None)

    # richValueRel index -> r:id
    rel_xml = ET.fromstring(zf.read("xl/richData/richValueRel.xml"))
    relindex_to_rid = [rel.get(f"{{{NS_R}}}id") for rel in rel_xml.findall(f"{{{NS_RVREL}}}rel")]

    # r:id -> media target path
    rels_xml = ET.fromstring(zf.read("xl/richData/_rels/richValueRel.xml.rels"))
    rid_to_media = {
        rel.get("Id"): posixpath.normpath(posixpath.join("xl/richData", rel.get("Target")))
        for rel in rels_xml.findall(f"{{{NS_PKGREL}}}Relationship")
    }

    def resolve_media_path(vm):
        try:
            rv_index = bk_to_rv[vm - 1]
            rel_index = rv_to_relindex[rv_index]
            rid = relindex_to_rid[rel_index]
            return rid_to_media.get(rid)
        except (IndexError, TypeError):
            return None

    # find cells in the Photo column that carry a vm attribute
    sheet_xml = ET.fromstring(zf.read(sheet_xml_path))
    from openpyxl.utils import get_column_letter
    photo_col_letter = get_column_letter(photo_col_0idx + 1)

    row_to_path = {}
    os.makedirs(out_dir, exist_ok=True)
    for c in sheet_xml.iter(f"{{{NS_MAIN}}}c"):
        ref = c.get("r") or ""
        vm = c.get("vm")
        if not vm or not ref.startswith(photo_col_letter) or not ref[len(photo_col_letter):].isdigit():
            continue
        excel_row = int(ref[len(photo_col_letter):])
        media_path = resolve_media_path(int(vm))
        if not media_path or media_path not in zf.namelist():
            continue
        ext = os.path.splitext(media_path)[1] or ".png"
        out_path = os.path.join(out_dir, f"row_{excel_row}{ext}")
        with open(out_path, "wb") as f:
            f.write(zf.read(media_path))
        row_to_path[excel_row] = out_path
    return row_to_path


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "usage: extract_excel_products.py <xlsx> <outdir>"}))
        sys.exit(1)

    xlsx_path, out_dir = sys.argv[1], sys.argv[2]
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_index, header_row_num = build_header_index(ws)
    os.makedirs(out_dir, exist_ok=True)

    row_to_image = extract_classic_images(ws, out_dir)
    if "photo" in header_index:
        with zipfile.ZipFile(xlsx_path) as zf:
            sheet_xml_path = find_sheet_xml_path(zf, ws.title)
            rich_images = extract_rich_value_images(zf, sheet_xml_path, header_index["photo"], out_dir)
            row_to_image.update(rich_images)

    results = []
    for row in ws.iter_rows(min_row=header_row_num + 1):
        values = [c.value for c in row]
        if not any(v not in (None, "") for v in values):
            continue

        def get(field):
            idx = header_index.get(field)
            if idx is None or idx >= len(values):
                return ""
            val = values[idx]
            return "" if val is None else str(val).strip()

        sku = get("sku")
        if not sku:
            continue

        results.append({
            "row": row[0].row,
            "sku": sku,
            "categoryName": get("categoryName"),
            "subCategory": get("subCategory"),
            "settingType": get("settingType"),
            "description": get("description"),
            "metalId": get("metalId"),
            "metalGms": get("metalGms"),
            "metalCombined": get("metalCombined"),
            "grossGms": get("grossGms"),
            "saleTotal": get("saleTotal"),
            "diamondPcs": get("diamondPcs"),
            "shape": get("shape"),
            "certificate": get("certificate"),
            "certNo": get("certNo"),
            "imagePath": row_to_image.get(row[0].row),
        })

    print(json.dumps(results))


if __name__ == "__main__":
    main()
